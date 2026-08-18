"""Evaluate frozen forecasts against newly published ANP observations.

The Atlas target and ANP's national weekly summary are deliberately kept as
separate series.  The former is calculated from station-day microdata using the
local W-SUN convention; the latter is an official, rounded Sunday-to-Saturday
aggregate.  They can be compared, but must not be silently concatenated.
"""

from __future__ import annotations

import json
import math
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


def _iso_date(value: object) -> str:
    stamp = pd.Timestamp(value)
    if pd.isna(stamp):
        raise ValueError("ANP summary contains an invalid reference date")
    return stamp.date().isoformat()


def load_anp_national_s10_summary(path: Path) -> dict[str, Any]:
    """Read the official national S10 row without following workbook links."""

    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "BRASIL" not in workbook.sheetnames:
            raise ValueError("ANP summary workbook has no BRASIL sheet")
        sheet = workbook["BRASIL"]
        rows = sheet.iter_rows(values_only=True)
        header: list[str] | None = None
        records: list[dict[str, object]] = []
        for values in rows:
            cells = list(values)
            if header is None:
                if cells and str(cells[0]).strip().upper() == "DATA INICIAL":
                    header = [str(value).strip() if value is not None else "" for value in cells]
                continue
            records.append(dict(zip(header, cells, strict=False)))
    finally:
        workbook.close()

    if header is None:
        raise ValueError("ANP summary header was not found in the BRASIL sheet")
    matches = [
        row
        for row in records
        if str(row.get("PRODUTO", "")).strip().upper() == "OLEO DIESEL S10"
        and str(row.get("BRASIL", "")).strip().upper() == "BRASIL"
    ]
    if len(matches) != 1:
        raise ValueError(f"Expected one national Diesel S10 row, found {len(matches)}")

    row = matches[0]
    unit = str(row.get("UNIDADE DE MEDIDA", "")).strip()
    if unit.casefold() != "r$/l":
        raise ValueError(f"Unexpected ANP summary unit: {unit!r}")
    period_start = _iso_date(row["DATA INICIAL"])
    period_end = _iso_date(row["DATA FINAL"])
    if pd.Timestamp(period_end) - pd.Timestamp(period_start) != pd.Timedelta(days=6):
        raise ValueError("ANP summary is not a seven-day Sunday-to-Saturday period")

    result = {
        "source_file": path.as_posix(),
        "period_start": period_start,
        "period_end": period_end,
        "atlas_w_sun_target_date": (
            pd.Timestamp(period_end) + pd.Timedelta(days=1)
        ).date().isoformat(),
        "geography": "BRASIL",
        "product": "OLEO DIESEL S10",
        "station_count": int(row["NÚMERO DE POSTOS PESQUISADOS"]),
        "unit": "BRL/L",
        "price": float(row["PREÇO MÉDIO REVENDA"]),
        "standard_deviation": float(row["DESVIO PADRÃO REVENDA"]),
        "minimum": float(row["PREÇO MÍNIMO REVENDA"]),
        "maximum": float(row["PREÇO MÁXIMO REVENDA"]),
        "coefficient_of_variation": float(row["COEF DE VARIAÇÃO REVENDA"]),
    }
    numeric = [
        result["price"],
        result["standard_deviation"],
        result["minimum"],
        result["maximum"],
        result["coefficient_of_variation"],
    ]
    if not all(math.isfinite(float(value)) for value in numeric):
        raise ValueError("ANP summary contains a non-finite statistic")
    if not 1 <= result["price"] <= 20:
        raise ValueError("ANP summary price is outside the plausible BRL 1-20/L range")
    if result["station_count"] <= 0:
        raise ValueError("ANP summary station count must be positive")
    return result


def score_frozen_forecast(
    forecast: dict[str, Any], actual: float, *, status: str
) -> dict[str, Any]:
    """Return transparent point, interval and direction diagnostics."""

    point = float(forecast["point"])
    current = float(forecast["current_price"])
    actual = float(actual)
    error = point - actual
    predicted_change = point - current
    realized_change = actual - current
    direction_match = (
        (predicted_change > 0 and realized_change > 0)
        or (predicted_change < 0 and realized_change < 0)
        or (predicted_change == 0 and realized_change == 0)
    )
    return {
        "status": status,
        "origin_date": forecast["origin_date"],
        "target_date": forecast["target_date"],
        "horizon_days": int(forecast["horizon_days"]),
        "current_price": current,
        "forecast": point,
        "actual": actual,
        "error": error,
        "absolute_error": abs(error),
        "absolute_percentage_error_pct": abs(error) / actual * 100,
        "predicted_change_pct": (point / current - 1) * 100,
        "realized_change_pct": (actual / current - 1) * 100,
        "direction_match": direction_match,
        "p10": float(forecast["p10"]),
        "p90": float(forecast["p90"]),
        "interval_contains_actual": float(forecast["p10"]) <= actual <= float(forecast["p90"]),
    }


def build_release_evaluation(
    *,
    summary_path: Path,
    frozen_forecast_path: Path,
    market_path: Path,
    generated_at: datetime | None = None,
) -> dict[str, Any]:
    """Build definitive operational and provisional official-summary scores."""

    summary = load_anp_national_s10_summary(summary_path)
    forecasts = json.loads(frozen_forecast_path.read_text(encoding="utf-8"))
    if not isinstance(forecasts, list):
        raise ValueError("Frozen forecast artifact must contain a list")
    by_target = {str(item["target_date"]): item for item in forecasts}

    market = pd.read_csv(market_path)
    if not {"semana_fim", "preco_medio"}.issubset(market.columns):
        raise ValueError("Market gold file lacks semana_fim/preco_medio")
    market["semana_fim"] = pd.to_datetime(market["semana_fim"], errors="raise")
    market_actual = {
        row.semana_fim.date().isoformat(): float(row.preco_medio)
        for row in market[["semana_fim", "preco_medio"]].itertuples(index=False)
    }

    definitive: list[dict[str, Any]] = []
    for target_date, forecast in by_target.items():
        if target_date in market_actual:
            definitive.append(
                score_frozen_forecast(
                    forecast,
                    market_actual[target_date],
                    status="definitive_operational_target",
                )
            )

    mapped_target = summary["atlas_w_sun_target_date"]
    provisional = None
    if mapped_target in by_target:
        provisional = score_frozen_forecast(
            by_target[mapped_target],
            summary["price"],
            status="provisional_official_summary_rounded",
        )

    stamp = generated_at or datetime.now(UTC)
    return {
        "generated_at": stamp.astimezone(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "frozen_forecast_file": frozen_forecast_path.as_posix(),
        "operational_scores": definitive,
        "official_summary": summary,
        "provisional_official_summary_score": provisional,
        "methodology": {
            "atlas_target": "station-day microdata aggregated with local W-SUN convention",
            "official_summary": "rounded ANP national aggregate for a Sunday-to-Saturday period",
            "rule": "compare and report separately; never append the rounded summary to the operational target",
        },
    }


def write_release_evaluation(payload: dict[str, Any], path: Path) -> Path:
    """Persist the release evaluation as deterministic UTF-8 JSON."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path
